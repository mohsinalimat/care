import json

import frappe
import re
from io import BytesIO
import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from care.utils.xlsxutils import handle_html

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

# columns_heading = ["Item Code", "Item Name", "Required By", "Description", "Item Group", "Brand",
#                "Quantity", "Stock UOM", "Target Warehouse", "UOM", "UOM Conversion Factor",
#                "Stock Qty", "Actual Qty", "Completed Qty", "Received Qty", "Rate", "Amount",
#                "Project", "Cost Center", "Expense Account"]

columns_heading = ["Item Code", "Item Name", "Quantity", "Target Warehouse"]



@frappe.whitelist()
def make_xlsx(material_demand_lst, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook(write_only=True)

    number = 0
    for res in material_demand_lst:
        sheet_name = str(res)
        ws = wb.create_sheet(sheet_name, number)

        for i, column_width in enumerate(column_widths):
            if column_width:
                ws.column_dimensions[get_column_letter(i + 1)].width = column_width

        row1 = ws.row_dimensions[1]
        row1.font = Font(name='Calibri', bold=True)

        # data = frappe.db.sql("""select item_code, item_name, schedule_date, description, item_group,brand,
        #         qty, stock_uom, warehouse, uom, conversion_factor,
        #         stock_qty, actual_qty, ordered_qty, received_qty, rate, amount,
        #         project, cost_center, expense_account
        #         from `tabMaterial Demand Item`
        #         where parent = '{0}'""".format(res), as_list=True)
        #
        data = frappe.db.sql("""select item_code, item_name, qty, warehouse
                from `tabMaterial Demand Item`
                where parent = '{0}'                 
                order by brand, item_name""".format(res), as_list=True)

        ws.append(columns_heading)
        for row in data:
            clean_row = []
            for item in row:
                pass
                if isinstance(item, str) and (sheet_name not in ['Data Import Template', 'Data Export']):
                    value = handle_html(item)
                else:
                    value = item

                if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
                    # Remove illegal characters from the string
                    value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)

                clean_row.append(value)

            ws.append(clean_row)
        number += 1

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


@frappe.whitelist()
def build_xlsx_response(material_demand_lst):
    if material_demand_lst:
        m_lst = json.loads(str(material_demand_lst))
        xlsx_file = make_xlsx(m_lst)
        frappe.response['filename'] = 'Material Demands.xlsx'
        frappe.response['filecontent'] = xlsx_file.getvalue()
        frappe.response['type'] = 'binary'


@frappe.whitelist()
def make_xlsx_summary(purchase_request, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook(write_only=True)
    p_doc = frappe.get_doc("Purchase Request", purchase_request)
    supp_lst = ['##efef']
    for res in p_doc.suppliers:
        supp_lst.append(res.supplier)
    sheet_name = purchase_request
    ws = wb.create_sheet(purchase_request, 0)
    for i, column_width in enumerate(column_widths):
        if column_width:
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width

    row1 = ws.row_dimensions[1]
    row1.font = Font(name='Calibri', bold=True)
    data = frappe.db.sql("""select p1.item_code, p1.item_name, p1.brand,
            ifnull((select p2.supplier_part_no from `tabItem Supplier`as p2 where p2.parent = p1.item_code and supplier in {0} limit 1),"") as part_number, 
            sum(p1.pack_order_qty)            
            from `tabPurchase Request Item` as p1
            where p1.parent = '{1}'
            group by p1.item_code, p1.item_name, p1.brand 
            order by p1.brand, p1.item_name""".format(tuple(supp_lst), purchase_request), as_list=True)

    ws.append(["Purchase Request", purchase_request, "", "Date", p_doc.date.strftime("%d/%m/%Y")])
    ws.append(["Company", p_doc.company, "", "Required By", p_doc.required_by.strftime("%d/%m/%Y")])
    ws.append(["Suppliers", p_doc.supplier_name])
    ws.append([])
    ws.append(["-----------"])
    ws.append(["Item Code", "Item Name", "Brand", "Supplier Part No.", "Pack Order Qty"])
    for row in data:
        clean_row = []
        for item in row:
            pass
            if isinstance(item, str) and (sheet_name not in ['Data Import Template', 'Data Export']):
                value = handle_html(item)
            else:
                value = item

            if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
                # Remove illegal characters from the string
                value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)

            clean_row.append(value)
        ws.append(clean_row)
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file